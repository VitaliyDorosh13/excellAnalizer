from __future__ import annotations

import hashlib
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, NamedTuple, Optional, Sequence, Union

from .models import (
    CellPreview,
    ParseDocumentRequest,
    SheetPreview,
    TablePreview,
    WorkbookPreview,
)

SUPPORTED_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


class ParserDependencyError(RuntimeError):
    """Raised when an optional parser dependency is missing."""


class WorkbookParserError(RuntimeError):
    """Raised when a workbook cannot be parsed safely."""


class TableRegion(NamedTuple):
    top_row: int
    bottom_row: int
    left_col: int
    right_col: int
    non_empty_cell_count: int


class EffectiveCell(NamedTuple):
    value: object
    formula: Optional[str]
    data_type: Optional[str]
    comment: Optional[str]
    is_merged: bool
    merge_parent: Optional[str]
    merge_range: Optional[str]
    row_outline_level: int
    column_outline_level: int
    row_hidden: bool
    column_hidden: bool
    fill_color: Optional[str]
    font_color: Optional[str]


def preview_workbook(request: ParseDocumentRequest) -> WorkbookPreview:
    workbook_path = _validate_workbook_path(Path(request.document_path))
    workbook = _load_workbook(workbook_path)
    try:
        selected_sheet_names = set(request.sheet_names)
        sheets: list[SheetPreview] = []
        warnings: list[str] = []
        parsed_sheet_names: set[str] = set()

        for worksheet in workbook.worksheets:
            if selected_sheet_names and worksheet.title not in selected_sheet_names:
                continue

            is_hidden = worksheet.sheet_state != "visible"
            if is_hidden and not request.include_hidden_sheets:
                continue

            parsed_sheet_names.add(worksheet.title)

            max_row = worksheet.max_row or 0
            max_column = worksheet.max_column or 0
            preview_row_count = min(max_row, request.max_rows_per_sheet)
            preview_column_count = min(max_column, request.max_columns_per_sheet)
            truncated = max_row > preview_row_count or max_column > preview_column_count

            if truncated:
                warnings.append(
                    f"Sheet '{worksheet.title}' preview was truncated to "
                    f"{preview_row_count} rows and {preview_column_count} columns."
                )

            table_regions = _detect_table_regions(
                worksheet=worksheet,
                max_row=preview_row_count,
                max_column=preview_column_count,
                max_tables=request.max_tables_per_sheet,
            )
            tables = _build_table_previews(
                worksheet=worksheet,
                workbook_path=workbook_path,
                regions=table_regions,
            )

            if not tables:
                warnings.append(
                    f"Sheet '{worksheet.title}' did not produce distinct table regions; "
                    "falling back to a sheet-level preview."
                )
                tables = [
                    _build_fallback_table_preview(
                        worksheet=worksheet,
                        workbook_path=workbook_path,
                        preview_row_count=preview_row_count,
                        preview_column_count=preview_column_count,
                    )
                ]

            sheets.append(
                SheetPreview(
                    id=_sheet_id(workbook_path, worksheet.title),
                    name=worksheet.title,
                    hidden=is_hidden,
                    used_range=_build_range(max_row, max_column),
                    row_count=max_row,
                    column_count=max_column,
                    merged_ranges=[str(item) for item in worksheet.merged_cells.ranges],
                    truncated=truncated,
                    tables=tables,
                )
            )

        missing_sheet_names = sorted(selected_sheet_names - parsed_sheet_names)
        if missing_sheet_names:
            warnings.append(
                "Requested sheets were not found or were filtered out: "
                + ", ".join(missing_sheet_names)
            )

        return WorkbookPreview(
            document_id=_document_id(workbook_path),
            source_name=workbook_path.name,
            source_path=str(workbook_path),
            source_size_bytes=workbook_path.stat().st_size,
            sheet_count=len(workbook.sheetnames),
            imported_at=datetime.now(timezone.utc),
            sheets=sheets,
            warnings=warnings,
        )
    finally:
        workbook.close()


def _load_workbook(workbook_path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ParserDependencyError(
            "openpyxl is required to parse Excel workbooks. Install backend dependencies first."
        ) from exc

    try:
        return load_workbook(
            filename=workbook_path,
            read_only=False,
            data_only=False,
        )
    except Exception as exc:
        raise WorkbookParserError(f"Unable to parse workbook '{workbook_path.name}'.") from exc


def _validate_workbook_path(workbook_path: Path) -> Path:
    resolved_path = workbook_path.expanduser().resolve()

    if not resolved_path.exists():
        raise FileNotFoundError(f"Workbook not found: {resolved_path}")

    if not resolved_path.is_file():
        raise ValueError(f"Expected a file path, got: {resolved_path}")

    if resolved_path.suffix.lower() not in SUPPORTED_WORKBOOK_SUFFIXES:
        raise ValueError(
            "Unsupported workbook format. Supported extensions: "
            + ", ".join(sorted(SUPPORTED_WORKBOOK_SUFFIXES))
        )

    return resolved_path


def _collect_preview_rows(worksheet, row_iterable: Iterable[Iterable[object]]) -> list[list[CellPreview]]:
    preview_rows: list[list[CellPreview]] = []

    for row in row_iterable:
        preview_row: list[CellPreview] = []
        for cell in row:
            effective_cell = _effective_cell(worksheet, cell.row, cell.column)
            preview_row.append(
                CellPreview(
                    address=cell.coordinate,
                    value=_serialize_value(effective_cell.value),
                    formula=effective_cell.formula,
                    data_type=effective_cell.data_type,
                    comment=effective_cell.comment,
                    is_merged=effective_cell.is_merged,
                    merge_parent=effective_cell.merge_parent,
                    merge_range=effective_cell.merge_range,
                    row_outline_level=effective_cell.row_outline_level,
                    column_outline_level=effective_cell.column_outline_level,
                    row_hidden=effective_cell.row_hidden,
                    column_hidden=effective_cell.column_hidden,
                    fill_color=effective_cell.fill_color,
                    font_color=effective_cell.font_color,
                )
            )

        preview_rows.append(preview_row)

    return preview_rows


def _detect_table_regions(worksheet, max_row: int, max_column: int, max_tables: int) -> list[TableRegion]:
    if max_row <= 0 or max_column <= 0:
        return []

    non_empty_columns_by_row: dict[int, list[int]] = {}
    non_empty_rows: list[int] = []

    for row_index in range(1, max_row + 1):
        row_columns: list[int] = []
        for column_index in range(1, max_column + 1):
            if _has_meaningful_value(_effective_value(worksheet, row_index, column_index)):
                row_columns.append(column_index)

        if row_columns:
            non_empty_columns_by_row[row_index] = row_columns
            non_empty_rows.append(row_index)

    if not non_empty_rows:
        return []

    row_bands = _cluster_indices(non_empty_rows, max_gap=1)
    regions: list[TableRegion] = []

    for band_start, band_end in row_bands:
        candidate_columns = sorted(
            {
                column_index
                for row_index in range(band_start, band_end + 1)
                for column_index in non_empty_columns_by_row.get(row_index, [])
            }
        )
        column_bands = _cluster_indices(candidate_columns, max_gap=1)

        for column_start, column_end in column_bands:
            region = _build_region_from_band(
                worksheet=worksheet,
                band_start=band_start,
                band_end=band_end,
                column_start=column_start,
                column_end=column_end,
            )
            if region is None:
                continue

            regions.append(region)
            if len(regions) >= max_tables:
                return _sort_regions(regions)

    return _sort_regions(regions)


def _build_region_from_band(
    worksheet,
    band_start: int,
    band_end: int,
    column_start: int,
    column_end: int,
) -> Optional[TableRegion]:
    top_row: Optional[int] = None
    bottom_row: Optional[int] = None
    left_col: Optional[int] = None
    right_col: Optional[int] = None
    non_empty_cell_count = 0

    for row_index in range(band_start, band_end + 1):
        for column_index in range(column_start, column_end + 1):
            if not _has_meaningful_value(_effective_value(worksheet, row_index, column_index)):
                continue

            non_empty_cell_count += 1
            top_row = row_index if top_row is None else min(top_row, row_index)
            bottom_row = row_index if bottom_row is None else max(bottom_row, row_index)
            left_col = column_index if left_col is None else min(left_col, column_index)
            right_col = column_index if right_col is None else max(right_col, column_index)

    if (
        top_row is None
        or bottom_row is None
        or left_col is None
        or right_col is None
        or non_empty_cell_count == 0
    ):
        return None

    row_count = bottom_row - top_row + 1
    column_count = right_col - left_col + 1

    if row_count == 1 and column_count == 1:
        return None

    if non_empty_cell_count < 2:
        return None

    return TableRegion(
        top_row=top_row,
        bottom_row=bottom_row,
        left_col=left_col,
        right_col=right_col,
        non_empty_cell_count=non_empty_cell_count,
    )


def _build_table_previews(worksheet, workbook_path: Path, regions: Sequence[TableRegion]) -> list[TablePreview]:
    sheet_id = _sheet_id(workbook_path, worksheet.title)
    tables: list[TablePreview] = []

    for index, region in enumerate(regions, start=1):
        rows = _collect_preview_rows(
            worksheet,
            worksheet.iter_rows(
                min_row=region.top_row,
                max_row=region.bottom_row,
                min_col=region.left_col,
                max_col=region.right_col,
            )
        )
        tables.append(
            TablePreview(
                id=f"{sheet_id}:table-{index}",
                title=_detect_table_title(worksheet, region),
                range=_build_region_range(region),
                row_count=region.bottom_row - region.top_row + 1,
                column_count=region.right_col - region.left_col + 1,
                non_empty_cell_count=region.non_empty_cell_count,
                header_row_index=_detect_header_row_index(rows, region.top_row),
                rows=rows,
            )
        )

    return tables


def _build_fallback_table_preview(worksheet, workbook_path: Path, preview_row_count: int, preview_column_count: int) -> TablePreview:
    rows: list[list[CellPreview]] = []
    non_empty_cell_count = 0

    if preview_row_count > 0 and preview_column_count > 0:
        rows = _collect_preview_rows(
            worksheet,
            worksheet.iter_rows(
                min_row=1,
                max_row=preview_row_count,
                min_col=1,
                max_col=preview_column_count,
            )
        )
        non_empty_cell_count = sum(
            1
            for row in rows
            for cell in row
            if _has_meaningful_value(cell.value)
        )

    return TablePreview(
        id=f"{_sheet_id(workbook_path, worksheet.title)}:primary-preview",
        title="Primary sheet preview",
        range=_build_range(preview_row_count, preview_column_count),
        row_count=preview_row_count,
        column_count=preview_column_count,
        non_empty_cell_count=non_empty_cell_count,
        header_row_index=_detect_header_row_index(rows, 1),
        rows=rows,
    )


def _effective_cell(worksheet, row_index: int, column_index: int) -> EffectiveCell:
    cell = worksheet.cell(row=row_index, column=column_index)
    row_dimension = worksheet.row_dimensions[row_index]
    column_dimension = worksheet.column_dimensions[_column_name(column_index)]
    merged_range = _merged_range_for_cell(worksheet, row_index, column_index)
    if merged_range is None:
        return EffectiveCell(
            value=cell.value,
            formula=str(cell.value) if cell.data_type == "f" else None,
            data_type=cell.data_type,
            comment=cell.comment.text if cell.comment else None,
            is_merged=False,
            merge_parent=None,
            merge_range=None,
            row_outline_level=row_dimension.outlineLevel or 0,
            column_outline_level=column_dimension.outlineLevel or 0,
            row_hidden=bool(row_dimension.hidden),
            column_hidden=bool(column_dimension.hidden),
            fill_color=_cell_fill_color(cell),
            font_color=_style_color(cell.font.color),
        )

    parent_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return EffectiveCell(
        value=parent_cell.value,
        formula=str(parent_cell.value) if parent_cell.data_type == "f" else None,
        data_type=parent_cell.data_type,
        comment=parent_cell.comment.text if parent_cell.comment else None,
        is_merged=True,
        merge_parent=parent_cell.coordinate,
        merge_range=str(merged_range),
        row_outline_level=row_dimension.outlineLevel or 0,
        column_outline_level=column_dimension.outlineLevel or 0,
        row_hidden=bool(row_dimension.hidden),
        column_hidden=bool(column_dimension.hidden),
        fill_color=_cell_fill_color(parent_cell),
        font_color=_style_color(parent_cell.font.color),
    )


def _effective_value(worksheet, row_index: int, column_index: int) -> object:
    return _effective_cell(worksheet, row_index, column_index).value


def _merged_range_for_cell(worksheet, row_index: int, column_index: int):
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row_index <= merged_range.max_row
            and merged_range.min_col <= column_index <= merged_range.max_col
        ):
            return merged_range

    return None


def _cell_fill_color(cell) -> Optional[str]:
    if not cell.fill or not cell.fill.fill_type:
        return None

    return _style_color(cell.fill.fgColor)


def _style_color(color) -> Optional[str]:
    if color is None:
        return None

    color_type = getattr(color, "type", None)
    raw_rgb = getattr(color, "rgb", None)
    if color_type == "rgb" and isinstance(raw_rgb, str):
        normalized = raw_rgb.upper()
        if normalized in {"00000000", "00FFFFFF"}:
            return None
        return f"#{normalized[-6:]}"

    indexed = getattr(color, "indexed", None)
    if color_type == "indexed" and indexed is not None:
        if indexed == 64:
            return None
        return f"indexed:{indexed}"

    theme = getattr(color, "theme", None)
    if color_type == "theme" and theme is not None:
        tint = getattr(color, "tint", 0)
        if theme == 1 and tint == 0:
            return None
        return f"theme:{theme}:tint:{tint}"

    return None


def _serialize_value(value: object) -> Optional[Union[str, int, float, bool]]:
    if value is None:
        return None

    if isinstance(value, (str, int, float, bool)):
        return value

    if hasattr(value, "isoformat"):
        return value.isoformat()

    return str(value)


def _has_meaningful_value(value: object) -> bool:
    if value is None:
        return False

    if isinstance(value, str):
        return value.strip() != ""

    return True


def _cluster_indices(indices: Sequence[int], max_gap: int) -> list[tuple[int, int]]:
    if not indices:
        return []

    clusters: list[tuple[int, int]] = []
    start = indices[0]
    previous = indices[0]

    for index in indices[1:]:
        if index - previous <= max_gap + 1:
            previous = index
            continue

        clusters.append((start, previous))
        start = index
        previous = index

    clusters.append((start, previous))
    return clusters


def _build_region_range(region: TableRegion) -> str:
    return (
        f"{_column_name(region.left_col)}{region.top_row}:"
        f"{_column_name(region.right_col)}{region.bottom_row}"
    )


def _detect_table_title(worksheet, region: TableRegion) -> Optional[str]:
    title_row_index = region.top_row - 1
    if title_row_index < 1:
        return None

    candidate_values: list[str] = []
    seen_sources: set[str] = set()
    scan_start = max(1, region.left_col - 1)
    scan_end = min(worksheet.max_column or region.right_col, region.right_col + 1)

    for column_index in range(scan_start, scan_end + 1):
        effective_cell = _effective_cell(worksheet, title_row_index, column_index)
        source_address = effective_cell.merge_parent or worksheet.cell(row=title_row_index, column=column_index).coordinate
        if source_address in seen_sources:
            continue

        seen_sources.add(source_address)
        if not _has_meaningful_value(effective_cell.value):
            continue

        candidate = str(effective_cell.value).strip()
        if candidate:
            candidate_values.append(candidate)

    if len(candidate_values) != 1:
        return None

    title = candidate_values[0]
    if len(title) > 120:
        return None

    if re.search(r"\d{4,}", title):
        return None

    return title


def _detect_header_row_index(rows: Sequence[Sequence[CellPreview]], start_row: int) -> Optional[int]:
    if not rows:
        return None

    for offset, row in enumerate(rows):
        populated_values_by_source: dict[str, str] = {}
        for cell in row:
            if not _has_meaningful_value(cell.value):
                continue

            source_address = cell.merge_parent or cell.address
            populated_values_by_source[source_address] = str(cell.value).strip()

        populated_values = list(populated_values_by_source.values())
        if not populated_values:
            continue

        # A single merged cell spanning a row is usually a section title, not a table header.
        if len(populated_values) == 1 and len(row) > 1:
            continue

        alpha_like_values = [
            value
            for value in populated_values
            if any(character.isalpha() for character in value)
        ]

        if alpha_like_values and len(alpha_like_values) >= max(1, len(populated_values) // 2):
            return start_row + offset

    return None


def _sort_regions(regions: Sequence[TableRegion]) -> list[TableRegion]:
    return sorted(
        regions,
        key=lambda region: (
            region.top_row,
            region.left_col,
            region.bottom_row,
            region.right_col,
        ),
    )


def _build_range(row_count: int, column_count: int) -> str:
    if row_count <= 0 or column_count <= 0:
        return ""

    return f"A1:{_column_name(column_count)}{row_count}"


def _column_name(column_index: int) -> str:
    if column_index < 1:
        return "A"

    name = ""
    current = column_index

    while current > 0:
        current, remainder = divmod(current - 1, 26)
        name = chr(65 + remainder) + name

    return name


def _document_id(workbook_path: Path) -> str:
    digest = hashlib.sha256()
    stat = workbook_path.stat()
    digest.update(str(workbook_path).encode("utf-8"))
    digest.update(str(stat.st_mtime_ns).encode("utf-8"))
    digest.update(str(stat.st_size).encode("utf-8"))
    return digest.hexdigest()[:16]


def _sheet_id(workbook_path: Path, sheet_name: str) -> str:
    digest = hashlib.sha1()
    digest.update(str(workbook_path).encode("utf-8"))
    digest.update(sheet_name.encode("utf-8"))
    return digest.hexdigest()[:12]
